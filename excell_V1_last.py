import openpyxl
from openpyxl.styles import Font
import logging
import os, re
from google_trans_new import google_translator


logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL)
logging.debug("Loading... \n")

counter = 0


def searchAndModifyFile():
    # TODO поиск файла в корневой папке
    fileDir = os.getcwd()
    fileList = os.listdir(fileDir)
    fileSaved = os.listdir(os.path.join('SAM', fileDir))
    logging.debug("{}".format(fileSaved))
    logging.debug("{}".format(fileDir))
    logging.debug("{}".format(fileList))
    # Находит все подходящие файлы в родной папке
    for i in fileList:
        if i.endswith(r'.xlsx'):
            if i.startswith(r'SAM'):
                logging.debug("List current excell files: {}".format(i))
                workingArea(i)


def engToRu(fileEng):
    # TODO Находит в названии -Eng- и меняет на -Ru-
    langRegex = re.compile(r'-Eng-')
    fileRu = langRegex.sub("-Ru-", fileEng)
    logging.debug("Eng to Ru: {}".format(fileRu))
    return fileRu


def google_tr_eng_ru(sheet_wb, new_tb, ltr, num, color="FF0000"):
    # TODO гугл переводчик с англ на рус
    translator = google_translator()
    global counter
    counter += 1

    trl = translator.translate('{}'.format(sheet_wb['{}{}'.format(ltr, num)].value), lang_tgt='ru')
    s = sheet_wb['{}{}'.format(ltr, num)].font.size
    n = sheet_wb['{}{}'.format(ltr, num)].font.name

    new_tb['A{}'.format(counter)] = sheet_wb['{}{}'.format(ltr, num)].value
    sheet_wb['{}{}'.format(ltr, num)] = new_tb['B{}'.format(counter)] = trl
    sheet_wb['{}{}'.format(ltr, num)].font = Font(size=s, name=n, color=color)


def myMaxRow(page):
    # TODO модифицированный поиск конца текста перевода(max_row - пустые строки кол А)
    counter = 0
    for num in range(page.max_row, 1, -1):
        if page['A{}'.format(num)].value is None:
            counter += 1
        else:
            break
    return page.max_row - counter


def proc(sheet_wb, sheet_tb, new_tb, ltr, num):
    # TODO основная часть
    try:
        if sheet_wb['{}{}'.format(ltr, num)].value is None or \
                sheet_wb['{}{}'.format(ltr, num)].value[0] == '=' or \
                sheet_wb['{}{}'.format(ltr, num)].value.isdigit():
            pass
        else:
            ownDict = False
            for cell in range(1, sheet_tb.max_row):
                if sheet_wb['{}{}'.format(ltr, num)].value == sheet_tb['A{}'.format(cell)].value:
                    ownDict = True
                    s = sheet_wb['{}{}'.format(ltr, num)].font.size
                    n = sheet_wb['{}{}'.format(ltr, num)].font.name
                    sheet_wb['{}{}'.format(ltr, num)] = sheet_tb['B{}'.format(cell)].value
                    sheet_wb['{}{}'.format(ltr, num)].font = Font(color="0408f7", size=s, name=n)
            if not ownDict:
                google_tr_eng_ru(sheet_wb, new_tb, ltr, num)
    except Exception as e:
        logging.debug(" {}".format(e))


def workingArea(path):
    # TODO Рабочая область манипулирования с файлом
    wb = openpyxl.load_workbook(path)
    tb = openpyxl.load_workbook('SAM/S_Translate.xlsx')

    sheet_wb = wb['Proforma']
    sheet_tb = tb['Proforma']
    new_tb = tb.create_sheet('new')

    # Рабочая область:
    # -------------------------------------------------------------------------
    """ Процесс прохода по ячейкам и сравнение содержимого с таблицей перевода
    замененый текст выделить зеленым фоном
    если не будет совпадения перевод будет осуществлен через гугл переводчик 
    замененый текст выделить красным
    
    подключить базу данных для перевода с базы если такого нет использовать 
    гугл переводчик для обновления базы
    """
    m_m_r = myMaxRow(sheet_wb)

    # TODO выборка документа по частям
    for num in range(1, 14):
        for ltr in "ABCDEFGHIJKLMNO":
            proc(sheet_wb, sheet_tb, new_tb, ltr, num)
    for num in range(m_m_r - 47, m_m_r + 1):
        for ltr in "ABCDEFGHIJKLMNO":
            proc(sheet_wb, sheet_tb, new_tb, ltr, num)
    for num in range(14, m_m_r - 47):
        for ltr in "EGHIJKLN":
            proc(sheet_wb, sheet_tb, new_tb, ltr, num)
    # -------------------------------------------------------------------------

    wb.save('SAM/{}'.format(engToRu(path)))
    tb.save('SAM/S_Translate.xlsx')
    wb.close()
    tb.close()


searchAndModifyFile()

logging.debug("-----------------------------------------\n")
logging.debug("Successful finish \n")
